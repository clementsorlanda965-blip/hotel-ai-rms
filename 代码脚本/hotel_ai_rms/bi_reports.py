"""
bi_reports.py — 酒店BI报表引擎
生成 RevPAR/ADR/OCC/GOP 专业分析 Excel，含公式图表条件格式
可独立运行，也可被 Streamlit 驾驶舱调用
"""
from datetime import date, timedelta, datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

OUTPUT_DIR = Path(r"E:\工作AI\酒店管理\数据分析")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ═══════════════════════════════════════════════════════════════
# 行业基准
# ═══════════════════════════════════════════════════════════════

BENCHMARKS = {
    "OCC": {"优秀": 80, "达标": 65, "需改善": 50},
    "RevPAR": {"优秀": 500, "达标": 350, "需改善": 200},
    "GOP率": {"优秀": 40, "达标": 32, "需改善": 20},
    "人房比": {"优秀": 1.0, "达标": 1.2, "需改善": 1.5},
}

SEASON_FACTORS = {
    1: 0.65, 2: 0.70, 3: 0.80, 4: 1.30, 5: 1.40,
    6: 1.10, 7: 1.35, 8: 1.35, 9: 1.30, 10: 1.50,
    11: 0.90, 12: 0.75,
}


# ═══════════════════════════════════════════════════════════════
# 数据生成 / 导入
# ═══════════════════════════════════════════════════════════════

def generate_sample_data(
    days: int = 30,
    base_occ: float = 68,
    base_adr: float = 500,
    base_rooms: int = 120,
    start_date: date = None,
    seed: int = 42,
) -> pd.DataFrame:
    """生成符合酒店行业规律的模拟经营数据。"""
    rng = np.random.default_rng(seed)
    if start_date is None:
        start_date = date.today() - timedelta(days=days)

    rows = []
    for d in range(days):
        cur = start_date + timedelta(days=d)
        m = cur.month
        dow = cur.weekday()
        season = SEASON_FACTORS.get(m, 1.0)
        weekend_boost = 1.12 if dow >= 5 else 1.0
        noise = float(rng.normal(0, 0.05))

        occ = base_occ * season * weekend_boost * (1 + noise)
        occ = max(20, min(98, occ))

        adr = base_adr * season * (1 + noise * 0.6)
        adr = max(150, round(adr, 2))

        revpar = round(occ / 100 * adr, 2)
        room_sold = round(occ / 100 * base_rooms)

        room_rev = round(adr * room_sold, 2)
        fb_rev = round(room_rev * rng.uniform(0.15, 0.35), 2)
        other_rev = round(room_rev * rng.uniform(0.03, 0.10), 2)
        total_rev = room_rev + fb_rev + other_rev

        cost_rate = rng.uniform(0.50, 0.68)
        total_cost = round(total_rev * cost_rate, 2)
        gop = round(total_rev - total_cost, 2)
        gop_rate = round(gop / total_rev * 100, 1) if total_rev > 0 else 0

        rows.append({
            "日期": cur,
            "OCC": round(occ, 1),
            "ADR": adr,
            "RevPAR": revpar,
            "客房收入": room_rev,
            "餐饮收入": fb_rev,
            "其他收入": other_rev,
            "总收入": total_rev,
            "总成本": total_cost,
            "GOP": gop,
            "GOP率": gop_rate,
            "已售房数": room_sold,
        })

    return pd.DataFrame(rows)


def load_from_database(start_date: str, end_date: str) -> pd.DataFrame:
    """从 SQLite 数据库加载实际经营数据。"""
    try:
        from database import get_daily_metrics
        df = get_daily_metrics(start_date, end_date)
        if df.empty:
            return None
        df["日期"] = pd.to_datetime(df["date"])
        df.rename(columns={
            "occ": "OCC", "adr": "ADR", "revpar": "RevPAR",
            "room_revenue": "客房收入", "fb_revenue": "餐饮收入",
            "other_revenue": "其他收入", "total_revenue": "总收入",
            "total_cost": "总成本", "gop": "GOP", "gop_rate": "GOP率",
            "room_sold": "已售房数",
        }, inplace=True)
        return df
    except ImportError:
        return None


# ═══════════════════════════════════════════════════════════════
# KPI 计算
# ═══════════════════════════════════════════════════════════════

def compute_kpi_summary(df: pd.DataFrame) -> dict:
    """从经营数据计算核心KPI汇总。"""
    total_days = len(df)
    avg_occ = df["OCC"].mean()
    avg_adr = df["ADR"].mean()
    avg_revpar = df["RevPAR"].mean()
    total_revenue = df["总收入"].sum()
    total_gop = df["GOP"].sum()
    avg_gop_rate = total_gop / total_revenue * 100 if total_revenue > 0 else 0

    # 趋势
    if total_days >= 14:
        mid = total_days // 2
        prev_revpar = df["RevPAR"].iloc[:mid].mean()
        curr_revpar = df["RevPAR"].iloc[mid:].mean()
        revpar_trend = (curr_revpar - prev_revpar) / prev_revpar * 100 if prev_revpar else 0
    else:
        revpar_trend = 0

    return {
        "周期天数": total_days,
        "日期范围": f"{df['日期'].min().strftime('%m/%d')} - {df['日期'].max().strftime('%m/%d')}",
        "平均OCC": round(avg_occ, 1),
        "平均ADR": round(avg_adr, 0),
        "平均RevPAR": round(avg_revpar, 0),
        "总收入": round(total_revenue, 0),
        "GOP": round(total_gop, 0),
        "GOP率": round(avg_gop_rate, 1),
        "RevPAR趋势": round(revpar_trend, 1),
        "总客房数": df["已售房数"].sum(),
        "日均已售房": round(df["已售房数"].mean(), 0),
    }


def rate_kpi(kpi_name: str, value: float) -> str:
    """行业基准对标评级。"""
    if kpi_name not in BENCHMARKS:
        return "—"
    b = BENCHMARKS[kpi_name]
    if value >= b["优秀"]:
        return "优秀"
    elif value >= b["达标"]:
        return "达标"
    else:
        return "需改善"


# ═══════════════════════════════════════════════════════════════
# 渠道分析
# ═══════════════════════════════════════════════════════════════

def generate_channel_analysis(total_rooms: int = 120, seed: int = 42) -> pd.DataFrame:
    """生成OTA渠道分析数据（模拟 + 支持导入真实数据）。"""
    rng = np.random.default_rng(seed)
    channels = [
        ("携程", 0.28, 0.15),
        ("美团", 0.18, 0.12),
        ("飞猪", 0.12, 0.10),
        ("官方小程序", 0.15, 0.03),
        ("抖音团购", 0.08, 0.05),
        ("Booking.com", 0.06, 0.15),
        ("前台散客", 0.07, 0.00),
        ("企业协议", 0.06, 0.00),
    ]

    rows = []
    for ch, share, commission in channels:
        rooms = round(total_rooms * 30 * share * rng.uniform(0.9, 1.1))
        adr = round(rng.uniform(380, 680), 0)
        revenue = rooms * adr
        comm_cost = revenue * commission

        rows.append({
            "渠道": ch,
            "间夜数": rooms,
            "占比": round(share * 100, 1),
            "平均ADR": adr,
            "收入": round(revenue, 0),
            "佣金率": f"{commission*100:.0f}%",
            "佣金成本": round(comm_cost, 0),
            "净收入": round(revenue - comm_cost, 0),
            "类型": "OTA" if commission > 0 else "直销",
        })

    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════
# Excel 报表生成
# ═══════════════════════════════════════════════════════════════

def generate_excel_report(
    df: pd.DataFrame,
    report_type: str = "月度经营报告",
    hotel_name: str = "我的酒店",
    output_path: str = None,
    budget_targets: dict = None,
    channel_df: pd.DataFrame = None,
) -> str:
    """生成专业酒店经营分析 Excel 报表。

    包含工作表：
    1. 经营概览 — KPI卡片 + 趋势图
    2. 每日明细 — 30天逐日数据
    3. 周趋势 — 按周汇总
    4. 渠道分析 — OTA/直销占比
    5. 预算对比 — 实际 vs 预算 (可选)
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        numbers,
    )
    from openpyxl.chart import LineChart, BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule

    wb = Workbook()

    # ── 颜色规范 ──
    BLUE = "1E4E8C"
    GOLD = "C8963E"
    DARK_BG = "0F2A4A"
    GREEN = "28A745"
    RED = "DC3545"
    ORANGE = "F0A040"
    LIGHT_GRAY = "F5F5F5"
    HEADER_BG = "1A3A6B"

    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="微软雅黑", size=16, bold=True, color=GOLD)
    kpi_value_font = Font(name="微软雅黑", size=22, bold=True, color=GOLD)
    kpi_label_font = Font(name="微软雅黑", size=10, color="8899BB")
    thin_border = Border(
        left=Side(style="thin", color="2A4A6D"),
        right=Side(style="thin", color="2A4A6D"),
        top=Side(style="thin", color="2A4A6D"),
        bottom=Side(style="thin", color="2A4A6D"),
    )

    kpi = compute_kpi_summary(df)

    # ═══════════════════════════════════════════════════════════
    # Sheet 1: 经营概览
    # ═══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "经营概览"

    # 标题
    ws1.merge_cells("A1:H1")
    ws1["A1"] = f"{hotel_name} — {report_type}"
    ws1["A1"].font = title_font
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:H2")
    ws1["A2"] = f"报告周期：{kpi['日期范围']} | 生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws1["A2"].font = Font(name="微软雅黑", size=9, color="8899BB")
    ws1["A2"].alignment = Alignment(horizontal="center")

    # KPI 卡片行
    kpi_items = [
        ("平均OCC", f"{kpi['平均OCC']}%", rate_kpi("OCC", kpi["平均OCC"])),
        ("平均ADR", f"¥{kpi['平均ADR']:.0f}", "—"),
        ("平均RevPAR", f"¥{kpi['平均RevPAR']:.0f}", rate_kpi("RevPAR", kpi["平均RevPAR"])),
        ("总收入", f"¥{kpi['总收入']:,.0f}", "—"),
        ("GOP率", f"{kpi['GOP率']}%", rate_kpi("GOP率", kpi["GOP率"])),
        ("RevPAR趋势", f"{kpi['RevPAR趋势']:+.1f}%", "—"),
    ]

    for i, (label, value, rating) in enumerate(kpi_items):
        col = i * 2 + 1
        ws1.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col + 1)
        ws1.cell(row=4, column=col, value=f"{label}\n{value}")
        ws1.cell(row=4, column=col).font = Font(name="微软雅黑", size=13, bold=True, color="FFFFFF")
        ws1.cell(row=4, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 评级色标
        if rating == "优秀":
            ws1.cell(row=6, column=col, value=f"★ {rating}")
            ws1.cell(row=6, column=col).font = Font(name="微软雅黑", size=9, bold=True, color=GREEN)
        elif rating == "达标":
            ws1.cell(row=6, column=col, value=f"● {rating}")
            ws1.cell(row=6, column=col).font = Font(name="微软雅黑", size=9, color=ORANGE)
        elif rating == "需改善":
            ws1.cell(row=6, column=col, value=f"▼ {rating}")
            ws1.cell(row=6, column=col).font = Font(name="微软雅黑", size=9, color=RED)

    ws1.row_dimensions[4].height = 50

    # RevPAR / OCC 趋势图
    chart_row = 8
    ws1.merge_cells(f"A{chart_row}:H{chart_row}")
    ws1[f"A{chart_row}"] = "RevPAR & OCC 趋势"
    ws1[f"A{chart_row}"].font = Font(name="微软雅黑", size=12, bold=True, color=GOLD)

    # 写入图表数据
    data_start = chart_row + 1
    ws1.cell(row=data_start, column=1, value="日期").font = header_font
    ws1.cell(row=data_start, column=1).fill = header_fill
    ws1.cell(row=data_start, column=2, value="RevPAR").font = header_font
    ws1.cell(row=data_start, column=2).fill = header_fill
    ws1.cell(row=data_start, column=3, value="OCC").font = header_font
    ws1.cell(row=data_start, column=3).fill = header_fill

    for i, (_, row) in enumerate(df.iterrows()):
        r = data_start + 1 + i
        ws1.cell(row=r, column=1, value=row["日期"].strftime("%m/%d"))
        ws1.cell(row=r, column=2, value=row["RevPAR"])
        ws1.cell(row=r, column=3, value=row["OCC"])

    n_rows = len(df)
    # RevPAR 折线
    chart1 = LineChart()
    chart1.title = "RevPAR 趋势"
    chart1.style = 10
    chart1.y_axis.title = "RevPAR (¥)"
    chart1.x_axis.title = "日期"
    data_ref1 = Reference(ws1, min_col=2, min_row=data_start, max_row=data_start + n_rows)
    cats1 = Reference(ws1, min_col=1, min_row=data_start + 1, max_row=data_start + n_rows)
    chart1.add_data(data_ref1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.width = 28
    chart1.height = 14
    from openpyxl.chart.series import DataPoint
    s = chart1.series[0]
    s.graphicalProperties.line.solidFill = GOLD
    ws1.add_chart(chart1, f"E{data_start}")

    # ═══════════════════════════════════════════════════════════
    # Sheet 2: 每日明细
    # ═══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("每日明细")

    headers = ["日期", "OCC(%)", "ADR(¥)", "RevPAR(¥)", "客房收入", "餐饮收入", "其他收入", "总收入", "总成本", "GOP", "GOP率(%)", "已售房数"]
    for j, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for i, (_, row) in enumerate(df.iterrows()):
        r = i + 2
        vals = [
            row["日期"].strftime("%Y-%m-%d"), row["OCC"], row["ADR"], row["RevPAR"],
            row["客房收入"], row["餐饮收入"], row["其他收入"], row["总收入"],
            row["总成本"], row["GOP"], row["GOP率"], row["已售房数"],
        ]
        for j, v in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=j, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # 条件格式: OCC
    ws2.conditional_formatting.add(
        f"B2:B{n_rows+1}",
        CellIsRule(operator="greaterThan", formula=["80"], fill=PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"), font=Font(color=GREEN)),
    )
    ws2.conditional_formatting.add(
        f"B2:B{n_rows+1}",
        CellIsRule(operator="lessThan", formula=["65"], fill=PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"), font=Font(color=RED)),
    )

    # 自动列宽
    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 14

    # ═══════════════════════════════════════════════════════════
    # Sheet 3: 周趋势
    # ═══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("周趋势")
    df["周"] = df["日期"].apply(lambda d: f"{d.isocalendar()[1]:02d}周")

    weekly = df.groupby("周").agg({
        "OCC": "mean", "ADR": "mean", "RevPAR": "mean",
        "总收入": "sum", "GOP": "sum", "已售房数": "sum",
    }).reset_index()
    weekly["GOP率"] = round(weekly["GOP"] / weekly["总收入"] * 100, 1)

    wh = ["周", "平均OCC(%)", "平均ADR(¥)", "平均RevPAR(¥)", "总收入", "GOP", "GOP率(%)", "总间夜"]
    for j, h in enumerate(wh, 1):
        cell = ws3.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for i, (_, row) in enumerate(weekly.iterrows()):
        vals = [row["周"], round(row["OCC"], 1), round(row["ADR"], 0), round(row["RevPAR"], 0),
                round(row["总收入"], 0), round(row["GOP"], 0), round(row["GOP率"], 1), int(row["已售房数"])]
        for j, v in enumerate(vals, 1):
            ws3.cell(row=i + 2, column=j, value=v).border = thin_border

    # ═══════════════════════════════════════════════════════════
    # Sheet 4: 渠道分析
    # ═══════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("渠道分析")

    if channel_df is None:
        channel_df = generate_channel_analysis()

    ch_headers = ["渠道", "间夜数", "占比(%)", "平均ADR(¥)", "收入", "佣金率", "佣金成本", "净收入", "类型"]
    for j, h in enumerate(ch_headers, 1):
        cell = ws4.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for i, (_, row) in enumerate(channel_df.iterrows()):
        vals = [row["渠道"], row["间夜数"], row["占比"], row["平均ADR"], row["收入"],
                row["佣金率"], row["佣金成本"], row["净收入"], row["类型"]]
        for j, v in enumerate(vals, 1):
            ws4.cell(row=i + 2, column=j, value=v).border = thin_border

    # 渠道收入饼图
    pie = PieChart()
    pie.title = "渠道收入占比"
    pie_data = Reference(ws4, min_col=5, min_row=1, max_row=1 + len(channel_df))
    pie_cats = Reference(ws4, min_col=1, min_row=2, max_row=1 + len(channel_df))
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    pie.width = 18
    pie.height = 12
    ws4.add_chart(pie, "A12")

    # ═══════════════════════════════════════════════════════════
    # Sheet 5: 预算对比 (可选)
    # ═══════════════════════════════════════════════════════════
    if budget_targets:
        ws5 = wb.create_sheet("预算对比")
        bh = ["指标", "实际", "预算", "差异", "达成率"]
        for j, h in enumerate(bh, 1):
            cell = ws5.cell(row=1, column=j, value=h)
            cell.font = header_font
            cell.fill = header_fill

        comp_items = [
            ("OCC(%)", kpi["平均OCC"], budget_targets.get("occ", 0)),
            ("ADR(¥)", kpi["平均ADR"], budget_targets.get("adr", 0)),
            ("RevPAR(¥)", kpi["平均RevPAR"], budget_targets.get("revpar", 0)),
            ("总收入(¥)", kpi["总收入"], budget_targets.get("revenue", 0)),
            ("GOP率(%)", kpi["GOP率"], budget_targets.get("gop_rate", 0)),
        ]
        for i, (label, actual, budget) in enumerate(comp_items):
            diff = actual - budget
            rate = actual / budget * 100 if budget else 0
            ws5.cell(row=i + 2, column=1, value=label)
            ws5.cell(row=i + 2, column=2, value=actual)
            ws5.cell(row=i + 2, column=3, value=budget)
            ws5.cell(row=i + 2, column=4, value=round(diff, 1))
            ws5.cell(row=i + 2, column=5, value=f"{rate:.1f}%")

            # 超预算标绿 未达标标红
            if rate >= 100:
                ws5.cell(row=i + 2, column=5).font = Font(color=GREEN, bold=True)
            else:
                ws5.cell(row=i + 2, column=5).font = Font(color=RED, bold=True)

    # ── 保存 ──
    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = str(OUTPUT_DIR / f"{hotel_name}_{report_type}_{ts}.xlsx")
    else:
        output_path = str(Path(output_path))

    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════════════════════
# 预算执行率追踪
# ═══════════════════════════════════════════════════════════════

def budget_vs_actual_report(
    df: pd.DataFrame,
    budget: dict,
    hotel_name: str = "我的酒店",
    output_path: str = None,
) -> str:
    """专项：预算执行率分析报表。"""
    return generate_excel_report(
        df, report_type="预算执行分析", hotel_name=hotel_name,
        output_path=output_path, budget_targets=budget,
    )


# ═══════════════════════════════════════════════════════════════
# GOP 深度分析
# ═══════════════════════════════════════════════════════════════

def gop_deep_dive(df: pd.DataFrame, hotel_name: str = "我的酒店", output_path: str = None) -> str:
    """专项：GOP 经营毛利深度分析。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()
    ws = wb.active
    ws.title = "GOP分析"

    GOLD = "C8963E"
    HEADER_BG = "1A3A6B"
    GREEN = "28A745"
    RED = "DC3545"

    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="2A4A6D"),
        right=Side(style="thin", color="2A4A6D"),
        top=Side(style="thin", color="2A4A6D"),
        bottom=Side(style="thin", color="2A4A6D"),
    )

    ws.merge_cells("A1:I1")
    ws["A1"] = f"{hotel_name} — GOP 经营毛利深度分析"
    ws["A1"].font = Font(name="微软雅黑", size=16, bold=True, color=GOLD)

    headers = ["日期", "总收入", "客房收入", "餐饮收入", "其他收入", "总成本", "GOP", "GOP率(%)", "成本率(%)"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for i, (_, row) in enumerate(df.iterrows()):
        r = i + 4
        cost_rate = round(row["总成本"] / row["总收入"] * 100, 1) if row["总收入"] > 0 else 0
        vals = [
            row["日期"].strftime("%m/%d"), row["总收入"], row["客房收入"],
            row["餐饮收入"], row["其他收入"], row["总成本"],
            row["GOP"], row["GOP率"], cost_rate,
        ]
        for j, v in enumerate(vals, 1):
            ws.cell(row=r, column=j, value=v).border = thin_border

    # GOP vs 成本柱状图
    chart = BarChart()
    chart.type = "col"
    chart.title = "收入 vs 成本 vs GOP"
    chart.style = 10
    n = len(df)
    data_ref = Reference(ws, min_col=2, max_col=7, min_row=3, max_row=3 + n)
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + n)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 28
    chart.height = 14
    ws.add_chart(chart, "A20")

    if output_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = str(OUTPUT_DIR / f"{hotel_name}_GOP分析_{ts}.xlsx")

    wb.save(str(output_path))
    return str(output_path)


# ═══════════════════════════════════════════════════════════════
# CLI 独立运行
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("生成酒店BI报表...")
    df = generate_sample_data(days=30, base_occ=72, base_adr=520)
    path = generate_excel_report(df, report_type="月度经营报告", hotel_name="九寨沟测试酒店")
    print(f"报表已生成：{path}")

    path2 = gop_deep_dive(df, hotel_name="九寨沟测试酒店")
    print(f"GOP分析已生成：{path2}")
